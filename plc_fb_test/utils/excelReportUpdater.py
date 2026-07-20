# -*- coding: utf-8 -*-
#
# © 2007-2008 Lenze Drive Systems GmbH. All rights reserved.
# © 2008-2021 Lenze Automation GmbH. All rights reserved.
#
#-----------------------------------------------------------------------------------------------------------------------
# PyTeF file header (for PyTeF Internal Code) V1.4-01
#-----------------------------------------------------------------------------------------------------------------------
# Disable linter warnings -> OTHER DISABLES ARE NOT RECOMMENDED, CHANGES ARE AT YOUR OWN RISK!
# pylint: disable=wildcard-import,unused-wildcard-import
#
#-----------------------------------------------------------------------------------------------------------------------
# PyTeF Internal Code docstring (must be first definition in file)
#-----------------------------------------------------------------------------------------------------------------------
"""
    Excel Report updater, which will update report with the test results
"""
#-----------------------------------------------------------------------------------------------------------------------
# PyTeF Internal Code imports (must be defined before code)
#-----------------------------------------------------------------------------------------------------------------------
import os
from typing import Tuple, Union, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tests.testlib.interface.patterns.Singleton import Singleton
#-----------------------------------------------------------------------------------------------------------------------
# SVN keyword section
#-----------------------------------------------------------------------------------------------------------------------
FILE_URL = "$HeadURL: $"[10:-2]  # noqa: E501
FILE_REV = "$Revision: $"[11:-2]
FILE_DATE = "$LastChangedDate: $"[18:-2]
FILE_AUTHOR = "$LastChangedBy: raina $"[16:-2]

#-----------------------------------------------------------------------------------------------------------------------
# PyTeF file header end
#-----------------------------------------------------------------------------------------------------------------------
class ReportExlParser(metaclass=Singleton):
    '''
    Class for reading and updating report data
    '''
    prefixCls = "[ReportExlParser]"
    TESTCASE_INDEX = 1
    JIRA_COL = 1
    TEST_DESCRIPTION_COL = 3
    START_ROW = 2
    MODE_COL = 4
    IMPLEMENTATION_COL = 5
    SESSION_COL = 6
    RESULT_COL = 7
    REMARK_COL = 8
    REPORT_HEADER_CELL = 'A1'

    PASS_TEXT = "Pass"
    FAIL_TEXT = "Fail"
    NOT_EXECUTED_TEXT = "Not Executed"

    #-------------------------------------------------------------------------------------------------------------------
    def __init__(self, reportFile: str, sheetName: str, logger: Any) -> None:
        '''
        Constructor for XLSX Report Reader
        '''
        # Opening workbook for reading xls file
        fullPath = os.path.join(os.getcwd(), reportFile)
        self.logger = logger
        self.reportFilePath = fullPath
        self.workbook = openpyxl.load_workbook(self.reportFilePath)
        self.sheetName = sheetName

        self.validateReportFile()
        self._reportData = {}
        self._loadReportData()

    #-------------------------------------------------------------------------------------------------------------------
    def loadWorkbook(self) -> Tuple[Workbook, Worksheet]:
        '''
        Load workbook and worksheet and return both
        '''
        workbook = openpyxl.load_workbook(self.reportFilePath)
        worksheet = workbook[self.sheetName]
        return workbook, worksheet

    #-------------------------------------------------------------------------------------------------------------------
    def saveAndCloseWorkbook(self, workbook: Workbook) -> None:
        '''
            Save and close workbook
        '''
        workbook.save(self.reportFilePath)
        workbook.close()

    #-------------------------------------------------------------------------------------------------------------------
    def _loadReportData(self) -> None:
        """
            Load Report Data from Report xlsx file
        """
        try:
            _, worksheet = self.loadWorkbook()

            for row_idx, row in enumerate(worksheet.iter_rows(min_row = ReportExlParser.START_ROW + 1,
                                                              min_col = ReportExlParser.JIRA_COL,
                                                              max_col = ReportExlParser.JIRA_COL + 8
                                                              ), start = ReportExlParser.TESTCASE_INDEX):

                jiraId = str(row[ReportExlParser.JIRA_COL].value)
                testDesc = row[ReportExlParser.TEST_DESCRIPTION_COL].value
                modeRow = row[ReportExlParser.MODE_COL].row
                implementationRow = row[ReportExlParser.IMPLEMENTATION_COL].row
                sessionRow = row[ReportExlParser.SESSION_COL].row
                resultRow = row[ReportExlParser.RESULT_COL].row
                remarkRow = row[ReportExlParser.REMARK_COL].row

                self._reportData[jiraId] = {}
                self._reportData[jiraId]['testcaseNum'] = row_idx
                self._reportData[jiraId]['testDesc'] = testDesc
                self._reportData[jiraId]['modeRow'] = modeRow
                self._reportData[jiraId]['implementationRow'] = implementationRow
                self._reportData[jiraId]['sessionRow'] = sessionRow
                self._reportData[jiraId]['resultRow'] = resultRow
                self._reportData[jiraId]['remarkRow'] = remarkRow

        except Exception:  # pylint: disable=broad-exception-caught
            self.logger.exception("[%s] Exception occurred while loading report data..!", type(self).prefixCls)

    #-------------------------------------------------------------------------------------------------------------------
    def getReportData(self) -> dict:
        '''
        Returns report data stored in a dictionary
        '''
        return self._reportData

    #-------------------------------------------------------------------------------------------------------------------
    def updateResult(self, jiraId: Union[int, str], result: Union[bool, str], remark: str = None, mode: str = None,
                     implementation: str = None, session: str = None) -> bool:
        '''
        Update result in report file for respective jira ids
        '''
        try:
            workbook, worksheet = self.loadWorkbook()

            if isinstance(jiraId, int):
                jiraId = str(jiraId)

            resCol = ReportExlParser.RESULT_COL
            resRow = self._reportData[jiraId]['resultRow']

            if isinstance(result, bool):
                if result:
                    result = ReportExlParser.PASS_TEXT
                else:
                    result = ReportExlParser.FAIL_TEXT
            else:
                result = ReportExlParser.NOT_EXECUTED_TEXT

            worksheet[chr(resCol + 65) + str(resRow)] = result

            remCol = ReportExlParser.REMARK_COL
            remRow = self._reportData[jiraId]['remarkRow']
            worksheet[chr(remCol + 65) + str(remRow)] = remark

            modeCol = ReportExlParser.MODE_COL
            modeRow = self._reportData[jiraId]['modeRow']
            worksheet[chr(modeCol + 65) + str(modeRow)] = mode

            impCol = ReportExlParser.IMPLEMENTATION_COL
            impRow = self._reportData[jiraId]['implementationRow']
            worksheet[chr(impCol + 65) + str(impRow)] = implementation

            sessCol = ReportExlParser.SESSION_COL
            sessRow = self._reportData[jiraId]['sessionRow']
            worksheet[chr(sessCol + 65) + str(sessRow)] = session

            self.saveAndCloseWorkbook(workbook = workbook)

        except IndexError:
            self.logger.warning("%s Jira id not found in helper sheet", type(self).prefixCls)
            self.logger.warning("%s Could not update result in report", type(self).prefixCls)

        except Exception:  # pylint: disable=broad-exception-caught
            self.logger.exception("[%s] Exception occurred while updating report excel", type(self).prefixCls)

    #-------------------------------------------------------------------------------------------------------------------
    def validateReportFile(self) -> bool:
        """
        Validate if report is present or not
        """
        try:
            self.logger.info("%s Validating Report File - %s", type(self).prefixCls, self.reportFilePath)
            if os.path.exists(self.reportFilePath) and os.path.isfile(self.reportFilePath):
                self.logger.info("%s Report file exists at location - %s", type(self).prefixCls, self.reportFilePath)
                return True

            self.logger.error("%s Report not found at location : %s", type(self).prefixCls, self.reportFilePath)
            return False
        except Exception:  # pylint: disable=broad-exception-caught
            self.logger.exception("[%s] Exception Occurred while validation of Report", type(self).prefixCls)
            return False

    #-------------------------------------------------------------------------------------------------------------------
    def getTestDescription(self, jiraId: Union[int, str]) -> str:
        """
        return TC description which was read from report
        """
        if isinstance(jiraId, int):
            jiraId = str(jiraId)
        return self.getReportData().get(jiraId).get('testDesc')

    #-------------------------------------------------------------------------------------------------------------------
    def getTestCaseNumber(self, jiraId: Union[int, str]) -> str:
        """
        return TC description which was read from report
        """
        if isinstance(jiraId, int):
            jiraId = str(jiraId)
        return self.getReportData().get(jiraId).get('testcaseNum')

    #-------------------------------------------------------------------------------------------------------------------
    def updateReportHeader(self, header: str) -> bool:
        """
        This method is used to update the header of the report file
        """
        try:
            workbook, worksheet = self.loadWorkbook()
            worksheet[ReportExlParser.REPORT_HEADER_CELL] = header
            self.saveAndCloseWorkbook(workbook = workbook)
            return True
        except Exception:  # pylint: disable=broad-exception-caught
            self.logger.exception("[%s] Exception occurred in _updateReportHeader()", type(self).prefixCls)
            return False
